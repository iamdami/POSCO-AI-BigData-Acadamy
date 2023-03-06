from openpyxl import Workbook

wb = Workbook() 
wb.sheetnames
ws = wb["Sheet"]
ws.title = "sheet1"

ws.append(["name", "age", "sex"])
ws.append(["Lee", 100, "male"])

ws.cell(3, 1).value = "Shin"
ws.cell(3, 2).value = 100
ws.cell(3, 3).value = "Female"

ws.cell(4, 1).value = "total"
ws.cell(4, 2).value = "=sum(B2:B3)"